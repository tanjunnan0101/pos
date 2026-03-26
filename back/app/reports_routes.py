"""
Sales / Revenue Reports API

Uses existing order and order_item data. Only paid/completed orders;
excludes removed and cancelled items. For restaurant owner revenue analysis.
"""

from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO, StringIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from . import models
from .db import get_session
from .permissions import Permission, require_permission
from .report_export_i18n import report_export_labels
from .security import get_current_user
from .work_session_serialization import serialize_work_session

router = APIRouter()

# Order statuses we count as "revenue"
REVENUE_STATUSES = {models.OrderStatus.paid, models.OrderStatus.completed}
# Item statuses we exclude from revenue
EXCLUDED_ITEM_STATUSES = {models.OrderItemStatus.cancelled}


def _revenue_date(order: models.Order) -> datetime | None:
    """Date used for attributing revenue (paid_at if set, else created_at)."""
    return order.paid_at or order.created_at


def _in_range(d: datetime | None, from_date: date, to_date: date) -> bool:
    if not d:
        return False
    d_date = d.date() if hasattr(d, "date") else d
    return from_date <= d_date <= to_date


def _get_revenue_items(
    session: Session,
    tenant_id: int,
    from_date: date,
    to_date: date,
):
    """Load orders and items that count toward revenue in the date range."""
    orders = session.exec(
        select(models.Order)
        .where(models.Order.tenant_id == tenant_id)
        .where(models.Order.deleted_at.is_(None))
        .where(models.Order.status.in_([s.value for s in REVENUE_STATUSES]))
        .order_by(models.Order.created_at.asc())
    ).all()

    result = []
    for order in orders:
        rev_date = _revenue_date(order)
        if not _in_range(rev_date, from_date, to_date):
            continue
        items = session.exec(
            select(models.OrderItem)
            .where(models.OrderItem.order_id == order.id)
            .where(models.OrderItem.removed_by_customer == False)
            .where(models.OrderItem.status != models.OrderItemStatus.cancelled)
        ).all()
        table = session.get(models.Table, order.table_id)
        waiter_id = None
        waiter_name = None
        if table:
            waiter_id = table.assigned_waiter_id
            if waiter_id is None and table.floor_id:
                floor = session.get(models.Floor, table.floor_id)
                if floor:
                    waiter_id = floor.default_waiter_id
            if waiter_id:
                u = session.get(models.User, waiter_id)
                waiter_name = (u.full_name or u.email) if u else str(waiter_id)
        table_name = table.name if table else "Unknown"
        for item in items:
            product = session.get(models.Product, item.product_id)
            category = (product.category or "Uncategorized") if product else "Uncategorized"
            subcategory = (product.subcategory or "") if product else ""
            unit_cost = getattr(item, "cost_cents", None) or 0
            revenue_cents = item.quantity * item.price_cents
            cost_cents = item.quantity * unit_cost
            result.append({
                "order_id": order.id,
                "date": rev_date,
                "table_id": order.table_id,
                "table_name": table_name,
                "waiter_id": waiter_id,
                "waiter_name": waiter_name or "Unassigned",
                "product_id": item.product_id,
                "product_name": item.product_name,
                "category": category,
                "subcategory": subcategory,
                "quantity": item.quantity,
                "price_cents": item.price_cents,
                "cost_cents": cost_cents,
                "revenue_cents": revenue_cents,
                "profit_cents": revenue_cents - cost_cents,
            })
    return result


def _build_report_payload(tenant_id: int, session: Session, from_date: date, to_date: date) -> dict:
    """Build full report dict for a tenant and date range."""
    if from_date > to_date:
        from_date, to_date = to_date, from_date
    rows = _get_revenue_items(session, tenant_id, from_date, to_date)

    # Summary by day
    by_day_agg: dict[str, dict] = defaultdict(
        lambda: {"revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "order_count": set()}
    )
    for r in rows:
        day = r["date"].strftime("%Y-%m-%d") if hasattr(r["date"], "strftime") else str(r["date"])[:10]
        by_day_agg[day]["revenue_cents"] += r["revenue_cents"]
        by_day_agg[day]["cost_cents"] += r["cost_cents"]
        by_day_agg[day]["profit_cents"] += r["profit_cents"]
        by_day_agg[day]["order_count"].add(r["order_id"])
    summary_daily = [
        {
            "date": d,
            "revenue_cents": data["revenue_cents"],
            "cost_cents": data["cost_cents"],
            "profit_cents": data["profit_cents"],
            "order_count": len(data["order_count"]),
        }
        for d, data in sorted(by_day_agg.items())
    ]
    total_revenue_cents = sum(r["revenue_cents"] for r in rows)
    total_cost_cents = sum(r["cost_cents"] for r in rows)
    total_profit_cents = total_revenue_cents - total_cost_cents
    total_orders = len(set(r["order_id"] for r in rows))

    # By product
    by_product: dict[tuple[int, str], dict] = defaultdict(
        lambda: {"quantity": 0, "revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "category": ""}
    )
    for r in rows:
        key = (r["product_id"], r["product_name"])
        by_product[key]["quantity"] += r["quantity"]
        by_product[key]["revenue_cents"] += r["revenue_cents"]
        by_product[key]["cost_cents"] += r["cost_cents"]
        by_product[key]["profit_cents"] += r["profit_cents"]
        if not by_product[key]["category"]:
            by_product[key]["category"] = r.get("category") or "Uncategorized"
    by_product_list = [
        {
            "product_id": k[0],
            "product_name": k[1],
            "category": v["category"],
            "quantity": v["quantity"],
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
        }
        for k, v in sorted(by_product.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    # By category
    by_category: dict[str, dict] = defaultdict(
        lambda: {"quantity": 0, "revenue_cents": 0, "cost_cents": 0, "profit_cents": 0}
    )
    for r in rows:
        c = r["category"] or "Uncategorized"
        by_category[c]["quantity"] += r["quantity"]
        by_category[c]["revenue_cents"] += r["revenue_cents"]
        by_category[c]["cost_cents"] += r["cost_cents"]
        by_category[c]["profit_cents"] += r["profit_cents"]
    by_category_list = [
        {
            "category": k,
            "quantity": v["quantity"],
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
        }
        for k, v in sorted(by_category.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    # By table
    by_table: dict[str, dict] = defaultdict(lambda: {"revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "order_count": set()})
    for r in rows:
        t = r["table_name"]
        by_table[t]["revenue_cents"] += r["revenue_cents"]
        by_table[t]["cost_cents"] += r["cost_cents"]
        by_table[t]["profit_cents"] += r["profit_cents"]
        by_table[t]["order_count"].add(r["order_id"])
    by_table_list = [
        {
            "table_name": k,
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
            "order_count": len(v["order_count"]),
        }
        for k, v in sorted(by_table.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    # By waiter
    by_waiter: dict[str, dict] = defaultdict(
        lambda: {"revenue_cents": 0, "cost_cents": 0, "profit_cents": 0, "order_count": set()}
    )
    for r in rows:
        w = r["waiter_name"]
        by_waiter[w]["revenue_cents"] += r["revenue_cents"]
        by_waiter[w]["cost_cents"] += r["cost_cents"]
        by_waiter[w]["profit_cents"] += r["profit_cents"]
        by_waiter[w]["order_count"].add(r["order_id"])
    by_waiter_list = [
        {
            "waiter_name": k,
            "revenue_cents": v["revenue_cents"],
            "cost_cents": v["cost_cents"],
            "profit_cents": v["profit_cents"],
            "order_count": len(v["order_count"]),
        }
        for k, v in sorted(by_waiter.items(), key=lambda x: -x[1]["revenue_cents"])
    ]

    # Reservations in date range (by reservation_date); source = public (token set) vs staff (no token); by status
    reservations = session.exec(
        select(models.Reservation)
        .where(models.Reservation.tenant_id == tenant_id)
        .where(models.Reservation.reservation_date >= from_date)
        .where(models.Reservation.reservation_date <= to_date)
    ).all()
    total_reservations = len(reservations)
    by_source: dict[str, int] = defaultdict(int)
    by_status: dict[str, int] = defaultdict(int)
    for r in reservations:
        source = "public" if r.token else "staff"
        by_source[source] += 1
        by_status[r.status.value] += 1

    # Overbooking: count slots (date, time) in range where reserved_guests > total_seats or reserved_parties > total_tables
    tables = session.exec(select(models.Table).where(models.Table.tenant_id == tenant_id)).all()
    total_seats = sum(t.seat_count for t in tables)
    total_tables = len(tables)
    active_reservations = [r for r in reservations if r.status in (models.ReservationStatus.booked, models.ReservationStatus.seated)]
    slot_aggregates: dict[tuple[date, time], tuple[int, int]] = defaultdict(lambda: (0, 0))  # (guests, parties)
    for r in active_reservations:
        key = (r.reservation_date, r.reservation_time)
        g, p = slot_aggregates[key]
        slot_aggregates[key] = (g + r.party_size, p + 1)
    overbooking_slots_count = sum(
        1 for (g, p) in slot_aggregates.values() if g > total_seats or p > total_tables
    )

    reservations_summary = {
        "total": total_reservations,
        "by_source": [
            {"source": k, "count": v} for k, v in sorted(by_source.items(), key=lambda x: -x[1])
        ],
        "by_status": [
            {"status": k, "count": v} for k, v in sorted(by_status.items(), key=lambda x: -x[1])
        ],
        "overbooking_slots_count": overbooking_slots_count,
    }

    average_revenue_per_order_cents = (
        total_revenue_cents // total_orders if total_orders else 0
    )
    return {
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "summary": {
            "total_revenue_cents": total_revenue_cents,
            "total_cost_cents": total_cost_cents,
            "total_profit_cents": total_profit_cents,
            "total_orders": total_orders,
            "average_revenue_per_order_cents": average_revenue_per_order_cents,
            "daily": summary_daily,
        },
        "reservations": reservations_summary,
        "by_product": by_product_list,
        "by_category": by_category_list,
        "by_table": by_table_list,
        "by_waiter": by_waiter_list,
    }


@router.get("/sales")
def get_sales_reports(
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
) -> dict:
    """Combined sales report for the given date range. Uses paid/completed orders only."""
    return _build_report_payload(current_user.tenant_id, session, from_date, to_date)


def _csv_stream(rows: list[dict], keys: list[str], header_row: list[str]) -> bytes:
    import csv

    # csv.writer requires a text stream; BytesIO expects bytes and raises TypeError.
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(header_row)
    for r in rows:
        writer.writerow([r.get(k, "") for k in keys])
    # utf-8-sig so Excel recognizes UTF-8 when headers contain non-ASCII (localized exports).
    return buf.getvalue().encode("utf-8-sig")


@router.get("/export")
def export_report(
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
    from_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    format: str = Query("csv", description="csv or xlsx"),
    report: str = Query("summary", description="summary, products, category, table, waiter"),
    lang: str | None = Query(None, description="UI language for headers (e.g. en, es, de)"),
) -> StreamingResponse:
    """Export report as CSV or Excel. Same date range as reports."""
    if from_date > to_date:
        from_date, to_date = to_date, from_date
    data = _build_report_payload(current_user.tenant_id, session, from_date, to_date)
    L = report_export_labels(lang)

    if format.lower() == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            from fastapi import HTTPException
            raise HTTPException(500, "Excel export requires openpyxl")
        wb = Workbook()
        # Summary sheet
        ws = wb.active
        ws.title = L["sheet_summary"][:31]
        ws.append(
            [
                L["date"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
                L["orders"],
            ]
        )
        for row in data["summary"]["daily"]:
            ws.append([
                row["date"],
                row["revenue_cents"],
                row.get("cost_cents", 0),
                row.get("profit_cents", 0),
                row["order_count"],
            ])
        ws.append([])
        s = data["summary"]
        ws.append(
            [
                L["total"],
                s["total_revenue_cents"],
                s.get("total_cost_cents", 0),
                s.get("total_profit_cents", 0),
                s["total_orders"],
            ]
        )
        # Reservations
        res = data.get("reservations", {})
        ws_res = wb.create_sheet(L["sheet_reservations"][:31])
        ws_res.append([L["source"], L["count"]])
        for row in res.get("by_source", []):
            sk = row["source"]
            src_label = L.get(f"source_{sk}", sk)
            ws_res.append([src_label, row["count"]])
        ws_res.append([])
        ws_res.append([L["status"], L["count"]])
        for row in res.get("by_status", []):
            st = row["status"]
            st_label = L.get(f"res_status_{st}", st)
            ws_res.append([st_label, row["count"]])
        ws_res.append([])
        ws_res.append([L["total"], res.get("total", 0)])
        # Products
        ws2 = wb.create_sheet(L["sheet_by_product"][:31])
        ws2.append(
            [
                L["product"],
                L["category"],
                L["quantity"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
            ]
        )
        for p in data["by_product"]:
            ws2.append([
                p["product_name"],
                p.get("category", ""),
                p["quantity"],
                p["revenue_cents"],
                p.get("cost_cents", 0),
                p.get("profit_cents", 0),
            ])
        # Category
        ws3 = wb.create_sheet(L["sheet_by_category"][:31])
        ws3.append(
            [
                L["category"],
                L["quantity"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
            ]
        )
        for c in data["by_category"]:
            ws3.append([
                c["category"],
                c["quantity"],
                c["revenue_cents"],
                c.get("cost_cents", 0),
                c.get("profit_cents", 0),
            ])
        # Table
        ws4 = wb.create_sheet(L["sheet_by_table"][:31])
        ws4.append(
            [
                L["table"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
                L["orders"],
            ]
        )
        for t in data["by_table"]:
            ws4.append([
                t["table_name"],
                t["revenue_cents"],
                t.get("cost_cents", 0),
                t.get("profit_cents", 0),
                t["order_count"],
            ])
        # Waiter
        ws5 = wb.create_sheet(L["sheet_by_waiter"][:31])
        ws5.append(
            [
                L["waiter"],
                L["revenue_cents"],
                L["cost_cents"],
                L["profit_cents"],
                L["orders"],
            ]
        )
        for w in data["by_waiter"]:
            ws5.append([
                w["waiter_name"],
                w["revenue_cents"],
                w.get("cost_cents", 0),
                w.get("profit_cents", 0),
                w["order_count"],
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=pos2-sales-{from_date}-{to_date}.xlsx"},
        )

    # CSV: single report type
    if report == "summary":
        rows = [
            {
                "date": r["date"],
                "revenue_cents": r["revenue_cents"],
                "cost_cents": r.get("cost_cents", 0),
                "profit_cents": r.get("profit_cents", 0),
                "order_count": r["order_count"],
            }
            for r in data["summary"]["daily"]
        ]
        keys = ["date", "revenue_cents", "cost_cents", "profit_cents", "order_count"]
        header_row = [
            L["date"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
            L["orders"],
        ]
    elif report == "products":
        rows = data["by_product"]
        keys = ["product_name", "category", "quantity", "revenue_cents", "cost_cents", "profit_cents"]
        header_row = [
            L["product"],
            L["category"],
            L["quantity"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
        ]
    elif report == "category":
        rows = data["by_category"]
        keys = ["category", "quantity", "revenue_cents", "cost_cents", "profit_cents"]
        header_row = [
            L["category"],
            L["quantity"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
        ]
    elif report == "table":
        rows = data["by_table"]
        keys = ["table_name", "revenue_cents", "cost_cents", "profit_cents", "order_count"]
        header_row = [
            L["table"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
            L["orders"],
        ]
    elif report == "waiter":
        rows = data["by_waiter"]
        keys = ["waiter_name", "revenue_cents", "cost_cents", "profit_cents", "order_count"]
        header_row = [
            L["waiter"],
            L["revenue_cents"],
            L["cost_cents"],
            L["profit_cents"],
            L["orders"],
        ]
    else:
        rows = data["summary"]["daily"]
        keys = ["date", "revenue_cents", "order_count"]
        header_row = [L["date"], L["revenue_cents"], L["orders"]]
    content = _csv_stream(rows, keys, header_row)
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=pos2-sales-{report}-{from_date}-{to_date}.csv"},
    )


@router.get("/work-sessions")
def report_work_sessions(
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
    from_date: str = Query(..., description="Start date YYYY-MM-DD (UTC day)"),
    to_date: str = Query(..., description="End date YYYY-MM-DD (UTC day, inclusive)"),
    user_id: int | None = Query(None, description="Filter by staff user id (optional)"),
) -> list[dict]:
    """All tenant staff clock-in/out rows in range (by started_at, UTC). Owner and admin only."""
    if current_user.tenant_id is None:
        raise HTTPException(status_code=403, detail="Tenant required")
    try:
        fd = datetime.strptime(from_date, "%Y-%m-%d").date()
        td = datetime.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format; use YYYY-MM-DD")
    if fd > td:
        raise HTTPException(status_code=400, detail="from_date must be <= to_date")
    start_utc = datetime.combine(fd, time.min, tzinfo=timezone.utc)
    end_exclusive = datetime.combine(td + timedelta(days=1), time.min, tzinfo=timezone.utc)
    q = (
        select(models.WorkSession)
        .where(models.WorkSession.tenant_id == current_user.tenant_id)
        .where(models.WorkSession.started_at >= start_utc)
        .where(models.WorkSession.started_at < end_exclusive)
    )
    if user_id is not None:
        q = q.where(models.WorkSession.user_id == user_id)
    rows = session.exec(q.order_by(models.WorkSession.started_at.desc())).all()
    out: list[dict] = []
    for ws in rows:
        u = session.get(models.User, ws.user_id)
        name = (u.full_name or u.email or "") if u else ""
        out.append(serialize_work_session(ws, name))
    return out
