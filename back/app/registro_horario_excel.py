"""
Monthly attendance Excel export.

Builds one sheet per employee with a fixed paper-like layout: company header,
daily grid (planned vs clocked, split morning/afternoon at 12:00 local), signature
column on every day row (cells left empty for ink), totals and footer text.

Aggregation uses the tenant IANA timezone when set; otherwise UTC.

Planned times come from Shift rows (working plan). Actual times come from
closed WorkSession rows, grouped by *local calendar date of session start*.
Each session’s net minutes (breaks excluded) count fully on that day — typical
for single same-day sessions; midnight-spanning sessions are rare.
"""

from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

from sqlmodel import Session, select

from . import models
from .work_session_serialization import work_session_net_duration_minutes

_SPANISH_WEEKDAY_ABBR = ("lun", "mar", "mié", "jue", "vie", "sáb", "dom")
_SPANISH_MONTHS = (
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
)


def _tenant_tz_name(tenant: models.Tenant) -> ZoneInfo:
    name = (tenant.timezone or "").strip()
    if name:
        try:
            return ZoneInfo(name)
        except (KeyError, ValueError):
            pass
    return ZoneInfo("UTC")


def _month_local_bounds(year: int, month: int, tz: ZoneInfo) -> tuple[date, date, datetime, datetime]:
    """First/last calendar day of month and UTC bounds [start_utc, end_utc) for session queries."""
    fd = date(year, month, 1)
    _, last = monthrange(year, month)
    ld = date(year, month, last)
    if month == 12:
        next_fd = date(year + 1, 1, 1)
    else:
        next_fd = date(year, month + 1, 1)
    start_utc = datetime.combine(fd, time.min, tzinfo=tz).astimezone(timezone.utc)
    end_utc = datetime.combine(next_fd, time.min, tzinfo=tz).astimezone(timezone.utc)
    return fd, ld, start_utc, end_utc


def _split_local_timerange_strings(
    day: date,
    t_start: time,
    t_end: time,
    tz: ZoneInfo,
) -> tuple[str, str]:
    """
    Split a same-day local time range into mañana / tarde strings (HH:MM–HH:MM).
    Split at 12:00 local wall clock.
    """
    a = datetime.combine(day, t_start, tzinfo=tz)
    b = datetime.combine(day, t_end, tzinfo=tz)
    noon = datetime.combine(day, time(12, 0), tzinfo=tz)
    if b <= a:
        return "", ""
    morning_parts: list[str] = []
    afternoon_parts: list[str] = []
    if a < noon:
        m_end = min(b, noon)
        morning_parts.append(f"{a.strftime('%H:%M')}–{m_end.strftime('%H:%M')}")
    if b > noon:
        a2 = max(a, noon)
        afternoon_parts.append(f"{a2.strftime('%H:%M')}–{b.strftime('%H:%M')}")
    return ("; ".join(morning_parts), "; ".join(afternoon_parts))


def _planned_ma_tarde_for_day(
    session: Session,
    tenant_id: int,
    user_id: int,
    day: date,
    tz: ZoneInfo,
) -> tuple[str, str]:
    rows = session.exec(
        select(models.Shift)
        .where(models.Shift.tenant_id == tenant_id)
        .where(models.Shift.user_id == user_id)
        .where(models.Shift.shift_date == day)
        .order_by(models.Shift.start_time.asc())
    ).all()
    m_all: list[str] = []
    t_all: list[str] = []
    for s in rows:
        if not s.start_time or not s.end_time:
            continue
        m, t = _split_local_timerange_strings(day, s.start_time, s.end_time, tz)
        if m:
            m_all.append(m)
        if t:
            t_all.append(t)
    return ("; ".join(m_all), "; ".join(t_all))


def _clip_utc_session_to_local_day(
    ws: models.WorkSession,
    tz: ZoneInfo,
    day: date,
) -> tuple[datetime, datetime] | None:
    """Intersect [started_at, ended_at] with the local civil day `day` in `tz`."""
    if ws.ended_at is None or ws.started_at is None:
        return None
    ls = ws.started_at.astimezone(tz)
    le = ws.ended_at.astimezone(tz)
    next_day = day + timedelta(days=1)
    day_start = datetime.combine(day, time.min, tzinfo=tz)
    day_end = datetime.combine(next_day, time.min, tzinfo=tz)
    seg_start = max(ls, day_start)
    seg_end = min(le, day_end)
    if seg_start >= seg_end:
        return None
    return seg_start, seg_end


def _segment_strings_ma_tarde(seg_start: datetime, seg_end: datetime, day: date, tz: ZoneInfo) -> tuple[str, str]:
    """Split an in-day local segment (aware datetimes) into mañana / tarde strings."""
    noon = datetime.combine(day, time(12, 0), tzinfo=tz)
    morning_parts: list[str] = []
    afternoon_parts: list[str] = []
    if seg_end <= seg_start:
        return "", ""
    if seg_start < noon:
        m_end = min(seg_end, noon)
        morning_parts.append(f"{seg_start.strftime('%H:%M')}–{m_end.strftime('%H:%M')}")
    if seg_end > noon:
        a_start = max(seg_start, noon)
        afternoon_parts.append(f"{a_start.strftime('%H:%M')}–{seg_end.strftime('%H:%M')}")
    return ("; ".join(morning_parts), "; ".join(afternoon_parts))


def _actual_ma_tarde_for_day(
    sessions_for_day: list[models.WorkSession],
    tz: ZoneInfo,
    day: date,
) -> tuple[str, str]:
    """
    Build display strings for clocked time, split at 12:00 local.
    Uses the intersection of each closed session with this calendar day.
    """
    morning_parts: list[str] = []
    afternoon_parts: list[str] = []
    epoch = datetime(1970, 1, 1, tzinfo=timezone.utc)

    def _key(ws: models.WorkSession) -> datetime:
        return ws.started_at if ws.started_at else epoch

    for ws in sorted(sessions_for_day, key=_key):
        clipped = _clip_utc_session_to_local_day(ws, tz, day)
        if clipped is None:
            continue
        seg_start, seg_end = clipped
        m, t = _segment_strings_ma_tarde(seg_start, seg_end, day, tz)
        if m:
            morning_parts.append(m)
        if t:
            afternoon_parts.append(t)
    return ("; ".join(morning_parts), "; ".join(afternoon_parts))


def _build_user_sheet(
    wb,
    session: Session,
    tenant: models.Tenant,
    user: models.User,
    year: int,
    month: int,
    tz: ZoneInfo,
    start_utc: datetime,
    end_utc: datetime,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(
        title=_sheet_title(user, year, month),
    )

    header_font = Font(bold=True)
    fill_hdr = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

    fd = date(year, month, 1)
    _, last_dom = monthrange(year, month)
    ld = date(year, month, last_dom)

    # Sessions in month for this user (query window covers TZ edges)
    stmt = (
        select(models.WorkSession)
        .where(models.WorkSession.tenant_id == tenant.id)
        .where(models.WorkSession.user_id == user.id)
        .where(models.WorkSession.started_at >= start_utc)
        .where(models.WorkSession.started_at < end_utc)
    )
    all_sessions = list(session.exec(stmt).all())

    # Group by local date of start
    by_day: dict[date, list[models.WorkSession]] = {}
    total_net_minutes = 0
    for wsess in all_sessions:
        if wsess.started_at is None:
            continue
        d = wsess.started_at.astimezone(tz).date()
        if d < fd or d > ld:
            continue
        by_day.setdefault(d, []).append(wsess)
        nm = work_session_net_duration_minutes(wsess, session)
        if nm is not None:
            total_net_minutes += nm

    cif = (tenant.cif or tenant.tax_id or "").strip()
    company = tenant.name or ""
    workplace = (tenant.address or "").strip()
    ccc = (tenant.ccc or "").strip()
    emp_num = (user.employee_number or "").strip()
    full_name = user.full_name or user.email or str(user.id)
    periodo = f"{_SPANISH_MONTHS[month - 1]} {year}"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "REGISTRO HORARIO"
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    row = 2
    for label, val in (
        ("Empresa", company),
        ("CIF / NIF", cif),
        ("Centro de trabajo", workplace),
        ("CCC (cuenta cotización)", ccc),
    ):
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=val or "")
        row += 1

    ws.cell(row=row, column=1, value="Empleado (número)").font = header_font
    ws.cell(row=row, column=2, value=emp_num)
    ws.cell(row=row, column=3, value="Nombre y apellidos").font = header_font
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws.cell(row=row, column=4, value=full_name)
    row += 1

    ws.cell(row=row, column=1, value="Periodo").font = header_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value=periodo)
    row += 1
    row += 1  # blank

    headers = [
        "Día",
        "Plan mañana",
        "Plan tarde",
        "Real mañana",
        "Real tarde",
        "Firma del empleado",
    ]
    for col_ix, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_ix, value=h)
        cell.font = header_font
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    header_row = row
    row += 1

    _, num_days = monthrange(year, month)
    for dnum in range(1, num_days + 1):
        day = date(year, month, dnum)
        wd = _SPANISH_WEEKDAY_ABBR[day.weekday()]
        day_label = f"{wd} {day.strftime('%d/%m/%Y')}"
        plan_m, plan_t = _planned_ma_tarde_for_day(session, tenant.id, user.id, day, tz)
        day_sessions = by_day.get(day, [])
        act_m, act_t = _actual_ma_tarde_for_day(day_sessions, tz, day)

        ws.cell(row=row, column=1, value=day_label)
        ws.cell(row=row, column=2, value=plan_m or "—")
        ws.cell(row=row, column=3, value=plan_t or "—")
        ws.cell(row=row, column=4, value=act_m or "—")
        ws.cell(row=row, column=5, value=act_t or "—")
        sig = ws.cell(row=row, column=6, value="")
        sig.alignment = Alignment(horizontal="center")
        for cix in range(1, 6):
            ws.cell(row=row, column=cix).alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    row += 1
    total_hours = total_net_minutes / 60.0
    ws.cell(row=row, column=1, value="Total horas netas (mes)").font = header_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    tot_cell = ws.cell(row=row, column=2, value=round(total_hours, 2))
    tot_cell.number_format = "0.00"
    row += 1

    ws.cell(row=row, column=1, value="Firma del empleado").font = header_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="")
    row += 1
    ws.cell(row=row, column=1, value="Firma de la empresa").font = header_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value="")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    legal = (
        "Documento elaborado conforme al artículo 34.9 ET y RD-ley 8/2019. "
        "Las firmas digitales no sustituyen a la firma manuscrita cuando la normativa aplicable la exija."
    )
    cleg = ws.cell(row=row, column=1, value=legal)
    cleg.alignment = Alignment(wrap_text=True)
    cleg.font = Font(italic=True, size=9)

    widths = (28, 18, 18, 18, 18, 22)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _sheet_title(user: models.User, year: int, month: int) -> str:
    base = user.full_name or user.email or str(user.id)
    raw = f"RH {month:02d}-{year} {base}"
    return raw[:31]


def build_registro_horario_xlsx_bytes(
    session: Session,
    tenant: models.Tenant,
    users: list[models.User],
    year: int,
    month: int,
) -> BytesIO:
    """Build workbook with one sheet per user. `users` must be non-empty."""
    if not users:
        raise ValueError("no users")

    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError("Excel export requires openpyxl") from e

    tz = _tenant_tz_name(tenant)
    _, _, start_utc, end_utc = _month_local_bounds(year, month, tz)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for user in users:
        _build_user_sheet(wb, session, tenant, user, year, month, tz, start_utc, end_utc)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
