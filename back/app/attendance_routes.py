"""
Employee Attendance Reports API (Legal-style Timesheet).
Generates monthly per-employee attendance in Excel format.
"""

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, col

from . import models
from .db import get_session
from .permissions import Permission, require_permission
from .work_session_serialization import _total_break_seconds

router = APIRouter()

@router.get("/attendance-excel")
def export_attendance_excel(
    current_user: Annotated[models.User, Depends(require_permission(Permission.REPORT_READ))],
    session: Session = Depends(get_session),
    year: int = Query(..., description="Year (YYYY)"),
    month: int = Query(..., ge=1, le=12, description="Month (1-12)"),
) -> StreamingResponse:
    """
    Export monthly per-employee attendance in Excel format.
    Columns: Date, Clock-In, Clock-Out, Break (min), Total Hours (net), Notes.
    """
    if current_user.tenant_id is None:
        raise HTTPException(status_code=403, detail="Tenant required")
    
    tenant_id = current_user.tenant_id

    # Define date range for the month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)

    # Query work sessions for this tenant and month
    # We'll use the start_date/end_date to filter sessions.
    # Note: A session might span across months, but we typically report by started_at.
    start_dt = datetime.combine(start_date, time.min, tzinfo=timezone.utc)
    end_dt = datetime.combine(end_date, time.max, tzinfo=timezone.utc)

    stmt = (
        select(models.WorkSession)
        .where(models.WorkSession.tenant_id == tenant_id)
        .where(models.WorkSession.started_at >= start_dt)
        .where(models.WorkSession.started_at <= end_dt)
        .order_by(models.WorkSession.started_at.asc())
    )
    sessions = session.exec(stmt).all()

    if not sessions:
        raise HTTPException(status_code=404, detail="No attendance records found for this period")

    # Group sessions by user
    user_sessions: dict[int, list[models.WorkSession]] = {}
    user_map: dict[int, models.User] = {}
    for ws in sessions:
        if ws.user_id not in user_sessions:
            user_sessions[ws.user_id] = []
        user_sessions[ws.user_id].append(ws)
        if ws.user_id not in user_map:
            user = session.get(models.User, ws.user_id)
            if user:
                user_map[ws.user_id] = user

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(500, "Excel export requires openpyxl")

    wb = Workbook()
    ws_sheet = wb.active
    ws_sheet.title = "Attendance"

    # Header row
    header = ["Employee Number", "Employee Name", "Date", "Clock-In", "Clock-Out", "Breaks (min)", "Net Hours", "Notes"]
    ws_sheet.append(header)
    for cell in ws_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Sort users by name for consistent report
    sorted_user_ids = sorted(user_sessions.keys(), key=lambda uid: (user_map[uid].full_name or "").lower())

    current_row = 2
    for uid in sorted_user_ids:
        user = user_map[uid]
        emp_num = user.employee_number or ""
        user_name = user.full_name or user.email or str(uid)
        
        # Sort sessions for this user by date
        user_ws = sorted(user_sessions[uid], key=lambda x: x.started_at)
        
        for sess in user_ws:
            # We'll report sessions that started in this month.
            # For a legal timesheet, we usually show each work session.
            
            started_at = sess.started_at
            ended_at = sess.ended_at
            
            # If session is still open, we use 'now' for the report (or just skip/mark as open)
            # For simplicity in a monthly report, we'll only report completed sessions or mark them.
            display_end = ended_at if ended_at else datetime.now(timezone.utc)
            
            # Calculate breaks
            break_sec = _total_break_seconds(session, sess, now_utc=display_end)
            break_min = break_sec // 60
            
            # Calculate net hours
            if sess.started_at and display_end:
                gross_sec = (display_end - sess.started_at).total_seconds()
                net_sec = max(0, gross_sec - break_sec)
                net_hours = net_sec / 3600
            else:
                net_hours = 0

            row = [
                emp_num if sess == user_ws[0] else "",  # Only show emp_num on first row of user's sessions
                user_name if sess == user_ws[0] else "",  # Only show name on first row of user's sessions
                sess.started_at.strftime("%Y-%m-%d"),
                sess.started_at.strftime("%H:%M"),
                display_end.strftime("%H:%M") if sess.ended_at else "OPEN",
                break_min,
                round(net_hours, 2),
                "",
            ]
            ws_sheet.append(row)
            current_row += 1

    # Style the sheet
    for row in range(2, current_row):
        # Center-align the numeric/date columns
        for col in range(3, 8):
            ws_sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"attendance_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
