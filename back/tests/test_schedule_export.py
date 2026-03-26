"""GET /schedule/export — per-worker monthly Excel export."""
from __future__ import annotations

import unittest
from datetime import date, time, timedelta
from io import BytesIO

from openpyxl import load_workbook
from pg_client_mixin import PgClientTestCase

from app import models, security


def _bearer_headers(user: models.User) -> dict[str, str]:
    data = {
        "sub": user.email,
        "tenant_id": user.tenant_id,
        "provider_id": getattr(user, "provider_id", None),
        "token_version": user.token_version,
    }
    token = security.create_access_token(data, expires_delta=timedelta(minutes=30))
    return {"Authorization": f"Bearer {token}"}


class TestScheduleExport(PgClientTestCase):
    def setUp(self) -> None:
        super().setUp()
        tenant = models.Tenant(name="Schedule export test")
        self.session.add(tenant)
        self.session.commit()
        self.session.refresh(tenant)
        self.tenant_id = tenant.id

        self.waiter = models.User(
            email="se-waiter@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Sara Waiter",
            tenant_id=self.tenant_id,
            role=models.UserRole.waiter,
        )
        self.session.add(self.waiter)
        self.session.commit()
        self.session.refresh(self.waiter)

        self.admin = models.User(
            email="se-admin@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Admin A",
            tenant_id=self.tenant_id,
            role=models.UserRole.admin,
        )
        self.session.add(self.admin)
        self.session.commit()
        self.session.refresh(self.admin)

        shift = models.Shift(
            tenant_id=self.tenant_id,
            user_id=self.waiter.id,
            shift_date=date(2025, 3, 15),
            start_time=time(9, 0),
            end_time=time(14, 0),
            label="Morning",
        )
        self.session.add(shift)
        self.session.commit()

    def test_export_xlsx_contains_shift(self) -> None:
        h = _bearer_headers(self.admin)
        r = self.client.get(
            "/schedule/export",
            params={"user_id": self.waiter.id, "year": 2025, "month": 3},
            headers=h,
        )
        self.assertEqual(r.status_code, 200, r.text)
        self.assertIn("spreadsheet", r.headers.get("content-type", ""))
        cd = r.headers.get("content-disposition") or ""
        self.assertIn("working-plan-", cd)
        self.assertIn(".xlsx", cd)
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(2, 1).value, "2025-03-15")
        self.assertEqual(ws.cell(2, 2).value, "09:00")
        self.assertEqual(ws.cell(2, 3).value, "14:00")
        self.assertEqual(ws.cell(2, 4).value, "Morning")
        self.assertEqual(ws.cell(2, 5).value, "Sara Waiter")

    def test_export_empty_month_header_only(self) -> None:
        h = _bearer_headers(self.admin)
        r = self.client.get(
            "/schedule/export",
            params={"user_id": self.waiter.id, "year": 2025, "month": 4},
            headers=h,
        )
        self.assertEqual(r.status_code, 200, r.text)
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 1)

    def test_export_user_not_in_tenant(self) -> None:
        other_tenant = models.Tenant(name="Other tenant SE")
        self.session.add(other_tenant)
        self.session.commit()
        self.session.refresh(other_tenant)
        outsider = models.User(
            email="se-outsider@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Outsider",
            tenant_id=other_tenant.id,
            role=models.UserRole.waiter,
        )
        self.session.add(outsider)
        self.session.commit()
        self.session.refresh(outsider)

        h = _bearer_headers(self.admin)
        r = self.client.get(
            "/schedule/export",
            params={"user_id": outsider.id, "year": 2025, "month": 3},
            headers=h,
        )
        self.assertEqual(r.status_code, 400, r.text)

    def test_plan_users_waiter_sees_schedulable_staff_without_user_read(self) -> None:
        """Waiter has schedule:read but not user:read; must still list plan workers for export dropdown."""
        h_waiter = _bearer_headers(self.waiter)
        r_users = self.client.get("/users", headers=h_waiter)
        self.assertEqual(r_users.status_code, 403, r_users.text)

        r = self.client.get("/schedule/plan-users", headers=h_waiter)
        self.assertEqual(r.status_code, 200, r.text)
        data = r.json()
        self.assertIsInstance(data, list)
        ids = {u["id"] for u in data}
        self.assertIn(self.waiter.id, ids)
        self.assertIn(self.admin.id, ids)

    def test_plan_users_admin_ok(self) -> None:
        h = _bearer_headers(self.admin)
        r = self.client.get("/schedule/plan-users", headers=h)
        self.assertEqual(r.status_code, 200, r.text)
        self.assertGreaterEqual(len(r.json()), 2)


if __name__ == "__main__":
    unittest.main()
