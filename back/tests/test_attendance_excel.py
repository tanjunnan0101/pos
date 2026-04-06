"""Monthly attendance Excel export (reports), including optional staff_ids filter."""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook
from pg_client_mixin import PgClientTestCase

from app import models, security


def _bearer_headers(user: models.User) -> dict[str, str]:
    data = {
        "sub": user.email,
        "tenant_id": user.tenant_id,
        "provider_id": getattr(user, "provider_id", None),
        "token_version": user.token_version,
    }
    token = security.create_access_token(data, expires_delta=timedelta(minutes=30))
    return {"Authorization": f"Bearer {token}"}


class TestAttendanceExcel(PgClientTestCase):
    def setUp(self) -> None:
        super().setUp()
        tenant = models.Tenant(name="Att Excel Test")
        self.session.add(tenant)
        self.session.commit()
        self.session.refresh(tenant)
        self.tenant_id = tenant.id

        self.staff_a = models.User(
            email="att-staff-a@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Staff Alpha",
            tenant_id=self.tenant_id,
            role=models.UserRole.waiter,
        )
        self.staff_b = models.User(
            email="att-staff-b@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Staff Beta",
            tenant_id=self.tenant_id,
            role=models.UserRole.waiter,
        )
        self.admin = models.User(
            email="att-admin@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Admin A",
            tenant_id=self.tenant_id,
            role=models.UserRole.admin,
        )
        self.session.add_all([self.staff_a, self.staff_b, self.admin])
        self.session.commit()
        self.session.refresh(self.staff_a)
        self.session.refresh(self.staff_b)
        self.session.refresh(self.admin)

        # March 2026 — deterministic month for assertions
        base = datetime(2026, 3, 15, 9, 0, tzinfo=timezone.utc)
        for user in (self.staff_a, self.staff_b):
            ws = models.WorkSession(
                tenant_id=self.tenant_id,
                user_id=user.id,
                started_at=base,
                ended_at=base + timedelta(hours=4),
            )
            self.session.add(ws)
        self.session.commit()

    def test_attendance_excel_with_staff_ids_returns_200_and_filters_rows(self) -> None:
        headers = _bearer_headers(self.admin)
        r = self.client.get(
            "/reports/attendance-excel",
            params={"year": 2026, "month": 3, "staff_ids": self.staff_a.id},
            headers=headers,
        )
        self.assertEqual(r.status_code, 200, r.text)
        self.assertIn("spreadsheetml", r.headers.get("content-type", ""))

        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        sheet = wb.active
        names_col_b = {
            row[1].value
            for row in sheet.iter_rows(min_row=2, max_col=2)
            if row[1].value
        }
        self.assertEqual(names_col_b, {"Staff Alpha"})

    def test_attendance_excel_invalid_staff_id_400(self) -> None:
        headers = _bearer_headers(self.admin)
        r = self.client.get(
            "/reports/attendance-excel",
            params={"year": 2026, "month": 3, "staff_ids": 999_999_999},
            headers=headers,
        )
        self.assertEqual(r.status_code, 400, r.text)
